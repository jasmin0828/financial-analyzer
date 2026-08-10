"""
财务分析核心逻辑 v2.0（支持多年趋势 + 图表）
- 数据提取（PDF / Excel），自动按年份归类
- 单年财务指标计算（30+ 指标）
- 多年趋势分析（YoY、CAGR、方向箭头）
- 多 Sheet Excel 报告（含 openpyxl 原生图表）
- 可独立调用，便于测试和扩展
"""

import re
import json
import traceback
from pathlib import Path
from datetime import datetime

# 工具版本号（遵循语义化版本 https://semver.org/）
# 主版本.次版本.修订号（例如 2.5.0）
# 变动规则：
#   - 主版本：不兼容的大改动
#   - 次版本：新增功能
#   - 修订号：bug 修复
__version__ = "2.5.0"

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


# ============================================================
# 财务科目标准化定义（带同义词映射）
# ============================================================

BS_ITEMS = {
    "货币资金": ["货币资金", "现金及现金等价物", "现金"],
    "应收账款": ["应收账款", "应收帐款", "应收账款合计"],
    "应收票据": ["应收票据", "应收票据及应收账款"],
    "存货": ["存货", "存货合计", "存贷"],
    "流动资产合计": ["流动资产合计", "流动资产总计", "流动资产"],
    "固定资产": ["固定资产", "固定资产净额", "固定资产合计"],
    "无形资产": ["无形资产", "无形资产净额"],
    "商誉": ["商誉"],
    "长期股权投资": ["长期股权投资", "长期股权投入"],
    "非流动资产合计": ["非流动资产合计", "非流动资产总计", "非流动资产"],
    "资产总计": ["资产总计", "资产合计", "总资产"],
    "短期借款": ["短期借款", "短期借款合计"],
    "应付账款": ["应付账款", "应付帐款"],
    "应付票据": ["应付票据"],
    "其他应付款": ["其他应付款", "其他应付款合计"],
    "一年内到期的非流动负债": ["一年内到期的非流动负债", "一年内到期非流动负债"],
    "流动负债合计": ["流动负债合计", "流动负债总计", "流动负债"],
    "长期借款": ["长期借款", "长期借款合计"],
    "应付债券": ["应付债券"],
    "非流动负债合计": ["非流动负债合计", "非流动负债总计"],
    "负债合计": ["负债合计", "负债总计", "总负债"],
    "实收资本": ["实收资本", "股本", "实收资本（或股本）"],
    "资本公积": ["资本公积"],
    "未分配利润": ["未分配利润"],
    "所有者权益合计": ["所有者权益合计", "所有者权益总计", "股东权益合计", "净资产",
                  "归属于母公司股东权益合计"],
    "少数股东权益": ["少数股东权益"],
    "负债和所有者权益总计": ["负债和所有者权益总计", "负债和所有者权益合计"],
}

IS_ITEMS = {
    "营业收入": ["营业收入", "营业总收入", "收入", "销售收入"],
    "营业成本": ["营业成本", "营业总成本", "销售成本", "成本"],
    "销售费用": ["销售费用"],
    "管理费用": ["管理费用"],
    "研发费用": ["研发费用", "研究开发费用", "R&D 费用"],
    "财务费用": ["财务费用"],
    "营业利润": ["营业利润"],
    "利润总额": ["利润总额", "税前利润"],
    "所得税费用": ["所得税费用", "所得税"],
    "净利润": ["净利润", "净利润合计"],
    "归母净利润": ["归属于母公司所有者的净利润", "归母净利润",
                  "归属母公司净利润", "归属于上市公司股东的净利润",
                  "归属于母公司"],
    "EPS": ["基本每股收益", "每股收益", "EPS", "基本 EPS"],
}

CF_ITEMS = {
    "经营活动产生的现金流量净额": ["经营活动产生的现金流量净额", "经营活动现金净额",
                                "经营现金净额", "经营活动现金流量净额"],
    "投资活动产生的现金流量净额": ["投资活动产生的现金流量净额", "投资活动现金净额"],
    "筹资活动产生的现金流量净额": ["筹资活动产生的现金流量净额", "筹资活动现金净额"],
    "现金及现金等价物净增加额": ["现金及现金等价物净增加额", "现金净增加额"],
    "购建固定资产、无形资产和其他长期资产支付的现金":
        ["购建固定资产、无形资产和其他长期资产支付的现金", "购建固定资产支付的现金", "资本支出"],
    "期末现金及现金等价物余额": ["期末现金及现金等价物余额", "现金及现金等价物期末余额"],
    "折旧与摊销": ["折旧与摊销", "折旧摊销"],
}


# ============================================================
# 年份识别
# ============================================================

def detect_year_from_filename(filepath):
    """从文件名识别年份（如 2023年报.pdf → 2023）

    支持格式：
    - 2023年报.pdf → 2023
    - 2025年12月利润表.xlsx → 2025
    - 2026年3月资产负债表.xlsx → 2026
    - annual_report_2025.pdf → 2025
    """
    name = Path(filepath).stem
    m = re.search(r'(20\d{2})', name)
    if m:
        return int(m.group(1))
    return None


def detect_period_from_filename(filepath):
    """从文件名识别期间（年份 + 月份）

    Returns:
        (year, month) 元组，如 (2025, 12) 或 (2026, 3)
        如果识别不出月份，返回 (year, 12)（默认年末）
    """
    name = Path(filepath).stem
    # 格式: 2025年12月 / 2026年3月 / 2025-12
    m = re.search(r'(20\d{2})年?[-_]?(\d{1,2})月?', name)
    if m:
        return int(m.group(1)), int(m.group(2))
    # 只有年份，默认为 12 月
    m = re.search(r'(20\d{2})', name)
    if m:
        return int(m.group(1)), 12
    return None, None


def detect_year_from_sheet(sheet_name):
    """从 sheet 名识别年份"""
    m = re.search(r'(20\d{2})', sheet_name)
    if m:
        return int(m.group(1))
    return None


def detect_period_from_sheet(sheet_name):
    """从 sheet 名识别期间（年份+月份）"""
    m = re.search(r'(20\d{2})年?[-_]?(\d{1,2})月?', sheet_name)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r'(20\d{2})', sheet_name)
    if m:
        return int(m.group(1)), 12
    return None, None


# ============================================================
# 数据提取
# ============================================================

def extract_from_excel(filepath, year=None):
    """从 Excel 财报中提取财务数据（智能识别布局）

    支持：
    - 单个 sheet 含 3 张表（按 sheet 名识别）
    - 多个 sheet 各含一张表（按 sheet 名识别）
    - A 股标准双栏布局资产负债表（资产左、负债右）
    - 智能根据内容识别 sheet 类型
    """
    result = {}
    try:
        xls = pd.ExcelFile(filepath, engine='openpyxl')
    except Exception:
        xls = pd.ExcelFile(filepath, engine='xlrd')

    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name,
                               engine='openpyxl', header=None, dtype=str)
        except Exception:
            continue

        # 智能识别 sheet 类型
        section_type = detect_sheet_type(sheet_name, df)
        if section_type is None:
            continue

        # 检测双栏布局（主要针对资产负债表）
        if section_type == "BS" and is_dual_column_layout(df):
            end_data, start_data = extract_dual_column_bs(df)
            extracted = end_data
            # 年初余额用 __prev__ 标记，后面由 analyze_files 移到上一年
            for k, v in start_data.items():
                extracted[f"__prev__{k}"] = v
        else:
            item_dict = {"BS": BS_ITEMS,
                         "IS": IS_ITEMS,
                         "CF": CF_ITEMS}[section_type]
            extracted = parse_table(df, item_dict)

        # 合并：保留已有数据，新数据只补充
        if section_type in result:
            for k, v in extracted.items():
                if k not in result[section_type]:
                    result[section_type][k] = v
        else:
            result[section_type] = extracted
    return result


def detect_sheet_type(sheet_name, df):
    """智能识别 sheet 是 BS/IS/CF 中的哪一种"""
    # 先看 sheet 名
    name_lower = sheet_name.lower()
    if any(kw in sheet_name for kw in ["资产负债表", "资负表", "BS", "Balance Sheet"]):
        return "BS"
    if any(kw in sheet_name for kw in ["利润表", "损益表", "收益表", "IS",
                                       "Income Statement", "Profit"]):
        return "IS"
    if any(kw in sheet_name for kw in ["现金流量表", "现金流", "CF", "Cash Flow"]):
        return "CF"

    # 看内容（前 10 行）
    try:
        head_text = ""
        for _, row in df.head(10).iterrows():
            for v in row.values:
                if v is not None and not (isinstance(v, float) and pd.isna(v)):
                    head_text += str(v) + " "
        # 资产负债表特征：同时有“资产”、“负债”
        if "资产" in head_text and "负债" in head_text and "所有者权益" in head_text:
            return "BS"
        if "营业收入" in head_text or "营业成本" in head_text:
            return "IS"
        if "经营活动" in head_text and "现金流量" in head_text:
            return "CF"
    except Exception:
        pass
    return None


def is_dual_column_layout(df):
    """检测是否是双栏布局（A 股标准资产负债表）"""
    for _, row in df.head(20).iterrows():
        row_str = " ".join([str(v) for v in row.values
                            if v is not None and not (isinstance(v, float) and pd.isna(v))])
        if "资产" in row_str and "负债" in row_str and ("期末" in row_str or "余额" in row_str):
            return True
    return False


def parse_value(v):
    """解析单元格值为数字"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    if not s or s in ("-", "--", "—", "－", ""):
        return None
    s_clean = s.replace(",", "").replace("，", "").replace(" ", "").replace("\u3000", "")
    s_clean = s_clean.replace("(", "-").replace(")", "")
    s_clean = s_clean.replace("（", "-").replace("）", "")
    try:
        return float(s_clean)
    except ValueError:
        return None


def extract_dual_column_bs(df):
    """从双栏布局的资产负债表中提取数据（同时获取期末 + 年初）

    双栏布局：A 股标准
    - 左侧：资产（列 A: 科目 / B: 行次 / C: 期末余额 / D: 年初余额）
    - 右侧：负债和所有者权益（列 E: 科目 / F: 行次 / G: 期末余额 / H: 年初余额）

    Returns:
        (end_data, start_data) 元组
        - end_data: {科目: 期末余额} （本期数）
        - start_data: {科目: 年初余额} （上年末数）
    """
    end_data = {}
    start_data = {}

    # 找表头行
    header_row_idx = None
    for i, row in df.iterrows():
        row_str = " ".join([str(v) for v in row.values
                            if v is not None and not (isinstance(v, float) and pd.isna(v))])
        if "资产" in row_str and "负债" in row_str and ("期末" in row_str or "余额" in row_str):
            header_row_idx = i
            break

    if header_row_idx is None:
        return end_data, start_data

    # 找列索引
    header_row = df.iloc[header_row_idx]
    asset_col = None       # 资产科目列
    asset_end_col = None   # 资产期末余额
    asset_start_col = None  # 资产年初余额
    liab_col = None         # 负债科目列
    liab_end_col = None     # 负债期末余额
    liab_start_col = None   # 负债年初余额

    for col_idx, val in header_row.items():
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        val_str = str(val).strip()
        # 资产列
        if val_str == "资产" or val_str.startswith("资产"):
            if asset_col is None:
                asset_col = col_idx
        # 负债列
        elif "负债" in val_str and "所有者权益" in val_str:
            if liab_col is None and asset_col is not None:
                liab_col = col_idx
        # 期末余额列
        elif "期末" in val_str:
            if asset_end_col is None:
                asset_end_col = col_idx
            elif liab_end_col is None:
                liab_end_col = col_idx
        # 年初余额列
        elif "年初" in val_str:
            if asset_start_col is None:
                asset_start_col = col_idx
            elif liab_start_col is None:
                liab_start_col = col_idx

    if asset_col is None or liab_col is None:
        return end_data, start_data

    # 默认列位置兑底（资产科后两列是期末/年初；负债同样）
    if asset_end_col is None:
        asset_end_col = asset_col + 2
    if liab_end_col is None:
        liab_end_col = liab_col + 2
    if asset_start_col is None:
        asset_start_col = asset_end_col + 1 if asset_end_col else asset_col + 3
    if liab_start_col is None:
        liab_start_col = liab_end_col + 1 if liab_end_col else liab_col + 3

    # 提取数据
    for i, row in df.iterrows():
        if i <= header_row_idx:
            continue

        # 提取资产
        try:
            asset_name = row.iloc[asset_col] if asset_col < len(row) else None
        except (IndexError, AttributeError):
            asset_name = None
        if asset_name and isinstance(asset_name, str):
            try:
                asset_end_value = row.iloc[asset_end_col] \
                    if asset_end_col < len(row) else None
            except (IndexError, AttributeError):
                asset_end_value = None
            try:
                asset_start_value = row.iloc[asset_start_col] \
                    if asset_start_col < len(row) else None
            except (IndexError, AttributeError):
                asset_start_value = None
            asset_end_value = parse_value(asset_end_value)
            asset_start_value = parse_value(asset_start_value)
            if asset_end_value is not None or asset_start_value is not None:
                for std_name, keywords in BS_ITEMS.items():
                    if any(kw in asset_name for kw in keywords):
                        if asset_end_value is not None:
                            end_data[std_name] = asset_end_value
                        if asset_start_value is not None:
                            start_data[std_name] = asset_start_value
                        break

        # 提取负债/所有者权益
        try:
            liab_name = row.iloc[liab_col] if liab_col < len(row) else None
        except (IndexError, AttributeError):
            liab_name = None
        if liab_name and isinstance(liab_name, str):
            try:
                liab_end_value = row.iloc[liab_end_col] \
                    if liab_end_col < len(row) else None
            except (IndexError, AttributeError):
                liab_end_value = None
            try:
                liab_start_value = row.iloc[liab_start_col] \
                    if liab_start_col < len(row) else None
            except (IndexError, AttributeError):
                liab_start_value = None
            liab_end_value = parse_value(liab_end_value)
            liab_start_value = parse_value(liab_start_value)
            if liab_end_value is not None or liab_start_value is not None:
                for std_name, keywords in BS_ITEMS.items():
                    if any(kw in liab_name for kw in keywords):
                        if liab_end_value is not None:
                            end_data[std_name] = liab_end_value
                        if liab_start_value is not None:
                            start_data[std_name] = liab_start_value
                        break

    return end_data, start_data


def parse_table(df, item_dict):
    """从表格中按关键词匹配提取数据（最长关键词优先原则）

    改进：避免 "流动资产合计" 被 "流动资产" 误匹配。
    策略：找所有匹配的行，选关键词最长的（即最具体的）。
    """
    extracted = {}
    for std_name, keywords in item_dict.items():
        best_row = None
        best_keyword = ""
        for _, row in df.iterrows():
            row_str = " ".join([str(x) for x in row.values
                                if pd.notna(x) and x is not None])
            if not row_str:
                continue
            # 在该行中找最长的匹配关键词
            for kw in keywords:
                if kw in row_str and len(kw) > len(best_keyword):
                    best_row = row
                    best_keyword = kw
        if best_row is not None:
            numbers = extract_numbers_from_row(list(best_row.values))
            if numbers:
                # 取绝对值最大的数字（业务金额通常最大）
                extracted[std_name] = max(numbers, key=abs)
    return extracted


def extract_numbers_from_row(values):
    """从一行单元格中提取数字（智能跳过行次列）"""
    numbers = []
    for v in values:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if not s or re.fullmatch(r"[\u4e00-\u9fa5\s%,()（）.\-]+", s):
            continue
        s_clean = s.replace(",", "").replace("，", "").replace(" ", "")
        s_clean = s_clean.replace("(", "-").replace(")", "")
        try:
            num = float(s_clean)
        except ValueError:
            continue
        # 跳过行次列（1-150 范围的纯整数通常不是金额）
        if 1 <= num <= 150 and num == int(num):
            continue
        numbers.append(num)
    return numbers


def extract_from_pdf(filepath):
    """从 PDF 财报中提取财务数据"""
    if pdfplumber is None:
        raise ImportError("请安装 pdfplumber: pip install pdfplumber")

    result = {}
    full_text = ""
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            full_text += text + "\n"
            tables = page.extract_tables() or []
            for table in tables:
                for row in table:
                    if row:
                        full_text += " ".join([str(c) for c in row if c]) + "\n"

    sections = split_pdf_sections(full_text)
    for section_name, section_text in sections.items():
        if section_name == "BS":
            result['BS'] = parse_text_data(section_text, BS_ITEMS)
        elif section_name == "IS":
            result['IS'] = parse_text_data(section_text, IS_ITEMS)
        elif section_name == "CF":
            result['CF'] = parse_text_data(section_text, CF_ITEMS)
    return result


def split_pdf_sections(text):
    """按章节切分 PDF 文本"""
    sections = {"BS": "", "IS": "", "CF": ""}
    bs_match = re.search(r"合并资产负债表", text) or re.search(r"资产负债表", text)
    is_match = re.search(r"合并利润表", text) or re.search(r"利润表", text)
    cf_match = re.search(r"合并现金流量表", text) or re.search(r"现金流量表", text)

    if bs_match and is_match:
        sections["BS"] = text[bs_match.start():is_match.start()]
    if is_match and cf_match:
        sections["IS"] = text[is_match.start():cf_match.start()]
    if cf_match:
        sections["CF"] = text[cf_match.start():]
    return sections


def parse_text_data(text, item_dict):
    """从纯文本中按关键词匹配提取数据"""
    extracted = {}
    lines = text.split("\n")
    for std_name, keywords in item_dict.items():
        for line in lines:
            if any(kw in line for kw in keywords):
                numbers = re.findall(r"-?\d{1,3}(?:,\d{3})*(?:\.\d+)?", line)
                if numbers:
                    nums = [float(n.replace(",", "")) for n in numbers]
                    extracted[std_name] = nums[0] if len(nums) == 1 else nums
                break
    return extracted


# ============================================================
# 指标计算（单年）
# ============================================================

def get_value(d, key, default=0):
    """从字典取值（支持 list）"""
    if not d:
        return default
    v = d.get(key)
    if isinstance(v, list):
        return v[0] if v else default
    return v if v is not None else default


def calculate_indicators(data, prev_data=None):
    """计算所有财务指标（单年数据）

    Args:
        data: 当年财报数据 {BS, IS, CF}
        prev_data: 上一年财报数据（可选，用于计算增长率类指标）
    """
    bs = data.get("BS", {})
    is_ = data.get("IS", {})
    cf = data.get("CF", {})

    revenue = get_value(is_, "营业收入")
    cost = get_value(is_, "营业成本")
    net_profit = get_value(is_, "净利润")
    parent_np = get_value(is_, "归母净利润") or net_profit
    operating_profit = get_value(is_, "营业利润")
    profit_total = get_value(is_, "利润总额")
    sell_exp = get_value(is_, "销售费用")
    admin_exp = get_value(is_, "管理费用")
    rd_expense = get_value(is_, "研发费用")
    fin_expense = get_value(is_, "财务费用")

    total_assets = get_value(bs, "资产总计")
    total_equity = get_value(bs, "所有者权益合计")
    minority_interest = get_value(bs, "少数股东权益")
    parent_equity = (total_equity - minority_interest
                     if total_equity and minority_interest else total_equity)
    total_liab = get_value(bs, "负债合计")
    current_assets = get_value(bs, "流动资产合计")
    current_liab = get_value(bs, "流动负债合计")
    inventory = get_value(bs, "存货")
    ar = get_value(bs, "应收账款")
    fixed_assets = get_value(bs, "固定资产")
    cash = get_value(bs, "货币资金")

    op_cf = get_value(cf, "经营活动产生的现金流量净额")
    capex = get_value(cf, "购建固定资产、无形资产和其他长期资产支付的现金")

    indicators = {}

    # ===== 盈利能力 =====
    if revenue and cost:
        indicators["毛利率"] = round((revenue - cost) / revenue * 100, 2)
    if revenue and net_profit:
        indicators["净利率"] = round(net_profit / revenue * 100, 2)
    if revenue and operating_profit:
        indicators["营业利润率"] = round(operating_profit / revenue * 100, 2)
    if revenue:
        indicators["销售费用率"] = round(sell_exp / revenue * 100, 2)
        indicators["管理费用率"] = round(admin_exp / revenue * 100, 2)
        indicators["研发费用率"] = round(rd_expense / revenue * 100, 2)
        indicators["财务费用率"] = round(fin_expense / revenue * 100, 2)
    if parent_equity and parent_np:
        indicators["ROE (归母净资产收益率)"] = round(parent_np / parent_equity * 100, 2)
    if total_assets and net_profit:
        indicators["ROA (总资产收益率)"] = round(net_profit / total_assets * 100, 2)
    invested_capital = total_equity + total_liab * 0.5
    if invested_capital > 0 and revenue and parent_np:
        indicators["ROIC (投入资本回报率, 估算)"] = round(
            parent_np / invested_capital * 100, 2)

    # ===== 偿债能力 =====
    if current_liab:
        indicators["流动比率"] = round(current_assets / current_liab, 2)
        indicators["速动比率"] = round((current_assets - inventory) / current_liab, 2)
        indicators["现金比率"] = round(cash / current_liab, 2)
    if total_assets:
        indicators["资产负债率"] = round(total_liab / total_assets * 100, 2)
    if total_equity:
        indicators["权益乘数"] = round(total_assets / total_equity, 2)
    long_cap = (total_equity + get_value(bs, "长期借款")
                + get_value(bs, "应付债券"))
    long_total = current_liab + get_value(bs, "非流动负债合计")
    if long_total > 0:
        indicators["长期资本适合率"] = round(long_cap / long_total, 2)

    # ===== 营运能力 =====
    if ar:
        indicators["应收账款周转率(次)"] = round(revenue / ar, 2)
        indicators["应收账款周转天数(天)"] = round(365 * ar / revenue, 1) if revenue else 0
    if inventory and cost:
        indicators["存货周转率(次)"] = round(cost / inventory, 2)
        indicators["存货周转天数(天)"] = round(365 * inventory / cost, 1) if cost else 0
    if total_assets:
        indicators["总资产周转率(次)"] = round(revenue / total_assets, 2)
    if fixed_assets:
        indicators["固定资产周转率(次)"] = round(revenue / fixed_assets, 2)
    if current_assets:
        indicators["流动资产周转率(次)"] = round(revenue / current_assets, 2)
    if total_assets and revenue:
        indicators["总资产周转天数(天)"] = round(365 * total_assets / revenue, 1)

    # ===== 现金流 =====
    if net_profit:
        indicators["经营现金流/净利润(现金含量)"] = round(op_cf / net_profit, 2)
    indicators["自由现金流(FCF)"] = round(op_cf - capex, 2)
    if revenue:
        indicators["经营现金流/营业收入"] = round(op_cf / revenue * 100, 2)
    if capex and revenue:
        indicators["资本支出/营业收入"] = round(capex / revenue * 100, 2)
    if total_assets:
        indicators["经营现金流/总资产"] = round(op_cf / total_assets * 100, 2)

    # ===== 杜邦分析 =====
    if revenue and total_assets and total_equity and net_profit:
        net_margin = net_profit / revenue * 100
        asset_turnover = revenue / total_assets
        equity_multiplier = total_assets / total_equity
        roe_dupont = net_margin * asset_turnover * equity_multiplier
        indicators["杜邦 ROE(%)"] = round(roe_dupont, 2)
        indicators["杜邦分解-净利率(%)"] = round(net_margin, 2)
        indicators["杜邦分解-总资产周转率"] = round(asset_turnover, 2)
        indicators["杜邦分解-权益乘数"] = round(equity_multiplier, 2)

    # ===== 收入与利润结构 =====
    if revenue and cost:
        indicators["毛利"] = round(revenue - cost, 2)
    if revenue and parent_np:
        indicators["归母净利润/营业收入"] = round(parent_np / revenue * 100, 2)

    # ===== 新增：利息保障倍数（偿债能力）=====
    # EBIT = 利润总额 + 财务费用（财务费用里主要是利息支出）
    if profit_total and fin_expense and fin_expense > 0:
        ebit = profit_total + fin_expense
        indicators["利息保障倍数"] = round(ebit / fin_expense, 2)

    # ===== 新增：EBITDA 利润率（盈利能力）=====
    # EBITDA = 营业利润 + 折旧与摊销
    depreciation = get_value(cf, "折旧与摊销")
    if operating_profit and revenue:
        ebitda = operating_profit + (depreciation or 0)
        indicators["EBITDA 利润率"] = round(ebitda / revenue * 100, 2)

    # ===== 新增：商誉/净资产（风险指标）=====
    goodwill = get_value(bs, "商誉")
    if goodwill is not None and total_equity:
        indicators["商誉/净资产"] = round(goodwill / total_equity * 100, 2)

    # ===== 新增：应收账款/营收（账期风险）=====
    if ar and revenue:
        indicators["应收账款/营收"] = round(ar / revenue * 100, 2)

    # ===== 新增：营业总收入增长率（发展能力，需要 prev_data）=====
    if prev_data:
        prev_is = prev_data.get("IS", {})
        prev_revenue = get_value(prev_is, "营业收入")
        prev_net_profit = get_value(prev_is, "净利润")
        prev_parent_np = get_value(prev_is, "归母净利润") or prev_net_profit
        if prev_revenue and revenue:
            indicators["营业总收入增长率"] = round(
                (revenue - prev_revenue) / prev_revenue * 100, 2)
        if prev_net_profit and net_profit:
            indicators["净利润增长率"] = round(
                (net_profit - prev_net_profit) / abs(prev_net_profit) * 100, 2)
        if prev_parent_np and parent_np:
            indicators["归母净利润增长率"] = round(
                (parent_np - prev_parent_np) / abs(prev_parent_np) * 100, 2)

    return indicators


# ============================================================
# 趋势分析（多年）
# ============================================================

def calculate_trends(indicators_by_year, years):
    """
    计算各指标的趋势（同比 YoY、CAGR 复合增长率、变化方向）

    Args:
        indicators_by_year: {year: {indicator: value}}，key 可以是 int 或 (year, month)
        years: 排序后的年份/期间列表

    Returns:
        {indicator: {
            "values": [(year, value), ...],
            "yoy": [{"year": y, "yoy": pct, "direction": "↑/↓/→"}, ...],
            "cagr": 复合增长率百分比 or None
        }}
    """
    if not years or len(years) < 1:
        return {}

    def get_year_part(key):
        """从 key 中提取年份部分（int 或 (int, int)）"""
        if isinstance(key, tuple):
            return key[0]
        return key

    trends = {}
    all_indicators = set()
    for ind in indicators_by_year.values():
        all_indicators.update(ind.keys())

    for indicator in all_indicators:
        values = []
        for year in years:
            v = indicators_by_year.get(year, {}).get(indicator)
            values.append((year, v))

        # 同比 YoY（跨期比较只看年份差）
        yoy_list = []
        for i in range(1, len(values)):
            prev_year, prev_v = values[i - 1]
            curr_year, curr_v = values[i]
            if prev_v is not None and curr_v is not None and prev_v != 0:
                yoy = (curr_v - prev_v) / abs(prev_v) * 100
                yoy_list.append({
                    "year": curr_year,
                    "yoy": round(yoy, 2),
                    "direction": "↑" if yoy > 0.5 else ("↓" if yoy < -0.5 else "→")
                })

        # CAGR（用年份差作 n）
        cagr = None
        first_year, first_v = values[0]
        last_year, last_v = values[-1]
        first_y = get_year_part(first_year)
        last_y = get_year_part(last_year)
        if (first_v is not None and last_v is not None
                and first_v > 0 and last_v > 0 and len(years) >= 2):
            n = last_y - first_y
            if n > 0:
                cagr_pow = pow(last_v / first_v, 1 / n)
                cagr = (cagr_pow - 1) * 100

        if yoy_list or cagr is not None:
            trends[indicator] = {
                "values": values,
                "yoy": yoy_list,
                "cagr": round(cagr, 2) if cagr is not None else None
            }

    return trends


# ============================================================
# 批量分析入口
# ============================================================

def analyze_files(file_list):
    """
    批量分析多个财报文件，自动按年份归类

    Args:
        file_list: 文件路径列表

    Returns:
        (data_by_year, indicators_by_year, years, trends, errors)
        data_by_year: {year: {BS, IS, CF}}
        indicators_by_year: {year: {indicator: value}}
        years: 排序后的年份列表
        trends: 趋势分析结果
        errors: [(filepath, error_msg), ...]
    """
    data_by_year = {}
    errors = []
    # 暂存 __prev__ 标记的“年初余额”数据，在合并后补充到上一年
    prev_pending = []

    for filepath in file_list:
        year, month = detect_period_from_filename(filepath)
        ext = Path(filepath).suffix.lower()
        try:
            if ext in ['.xlsx', '.xls']:
                data = extract_from_excel(filepath)
            elif ext == '.pdf':
                data = extract_from_pdf(filepath)
            else:
                errors.append((filepath, f"不支持的文件类型: {ext}"))
                continue

            bs_count = len(data.get("BS", {}))
            is_count = len(data.get("IS", {}))
            cf_count = len(data.get("CF", {}))

            if bs_count + is_count + cf_count == 0:
                errors.append((filepath, "未能提取到任何数据"))
                continue

            # 处理 __prev__ 标记的“年初余额”数据
            for section in ['BS', 'IS', 'CF']:
                if section in data:
                    prev_keys = [k for k in data[section]
                                 if k.startswith('__prev__')]
                    for k in prev_keys:
                        clean_k = k.replace('__prev__', '')
                        if year is not None:
                            prev_pending.append({
                                'section': section,
                                'key': clean_k,
                                'value': data[section][k],
                                'from_year': year,
                                'from_month': month,
                            })
                        # 从当前数据中删除（不当作本年数）
                        del data[section][k]

            # 如果识别不到年份，用当前年
            if year is None:
                year = datetime.now().year

            # 避免同年多份文件被覆盖：按 “年份-月份” 作为 key
            year_key = (year, month) if month else year
            if year_key in data_by_year:
                # 合并
                for section in ['BS', 'IS', 'CF']:
                    if section in data:
                        data_by_year[year_key].setdefault(
                            section, {}).update(data[section])
                # 记录每个 section 的源文件路径
                for section in ['BS', 'IS', 'CF']:
                    if section in data and data[section]:
                        data_by_year[year_key].setdefault(
                            '__files__', {})[section] = str(filepath)
                        break  # 一个文件对应一个 section
            else:
                data_by_year[year_key] = data
                # 记录每个 section 的源文件路径
                for section in ['BS', 'IS', 'CF']:
                    if section in data and data[section]:
                        data_by_year[year_key].setdefault(
                            '__files__', {})[section] = str(filepath)
                        break

        except Exception as e:
            errors.append((filepath, f"{str(e)}\n{traceback.format_exc()}"))

    # 补充“年初余额”到上一年末（仅在用户没提供的情况下）
    # 重要：年初余额总是上一年末（如 2026-3 BS 的年初 = 2025-12-31）
    for prev in prev_pending:
        # 年初余额是上一个资产负债表日，默认是上一年 12 月
        prev_year_key = (prev['from_year'] - 1, 12)
        if prev_year_key not in data_by_year:
            data_by_year[prev_year_key] = {}
        if prev['section'] not in data_by_year[prev_year_key]:
            data_by_year[prev_year_key][prev['section']] = {}
        # 仅在用户没提供该数据时补充
        if prev['key'] not in data_by_year[prev_year_key][prev['section']]:
            data_by_year[prev_year_key][prev['section']][prev['key']] = prev['value']

    # 年份排序（可能是 int 或 (int, int)）
    years = sorted(data_by_year.keys())

    # 标准化：把所有 list 值转为单个 float（取绝对值最大）
    for year_key, data in data_by_year.items():
        for section in ['BS', 'IS', 'CF']:
            if section in data:
                for k, v in list(data[section].items()):
                    if isinstance(v, list):
                        nums = [x for x in v if isinstance(x, (int, float))]
                        data[section][k] = max(nums, key=abs) if nums else None
    indicators_by_year = {}
    for year in years:
        # 上一年：可能跨年跳
        if isinstance(year, tuple):
            prev_year_key = (year[0] - 1, year[1])
        else:
            prev_year_key = year - 1
        prev_data = data_by_year.get(prev_year_key)
        indicators_by_year[year] = calculate_indicators(
            data_by_year[year], prev_data=prev_data)
    trends = calculate_trends(indicators_by_year, years)

    return data_by_year, indicators_by_year, years, trends, errors


# ============================================================
# 科目表提取（宽表 + 长表）
# ============================================================

# 报表类型中文映射
SECTION_LABELS = {
    "BS": ("资产负债表", "期末数"),
    "IS": ("利润表", "本期数"),
    "CF": ("现金流量表", "本期数"),
}


def extract_subjects_table(data_by_year, years):
    """从多年财报数据提取标准化科目表

    Returns:
        (wide_table, long_table):
        - wide_table: list of dict, 每条 {报表类型, 会计科目, 2023: v, 2024: v, ...}
        - long_table: list of dict, 每条 {年份, 报表类型, 会计科目, 类别, 数值}
    """
    # 收集所有 (报表类型, 科目) 并排序
    all_subjects = set()
    for year in years:
        data = data_by_year.get(year, {})
        for section in ["BS", "IS", "CF"]:
            for subject in data.get(section, {}).keys():
                all_subjects.add((section, subject))

    section_order = {"BS": 0, "IS": 1, "CF": 2}
    sorted_subjects = sorted(
        all_subjects,
        key=lambda x: (section_order.get(x[0], 99), x[1]))

    def year_label(y):
        """统一的年份字符串表示"""
        if isinstance(y, tuple):
            return f"{y[0]}-{str(y[1]).zfill(2)}"
        return str(y)

    # 宽表：行=科目，列=年份
    wide_table = []
    for section, subject in sorted_subjects:
        label, _ = SECTION_LABELS.get(section, (section, ""))
        row = {"报表类型": label, "会计科目": subject}
        for year in years:
            data = data_by_year.get(year, {})
            v = data.get(section, {}).get(subject)
            if isinstance(v, list):
                v = v[0] if v else None
            row[year_label(year)] = v
        wide_table.append(row)

    # 长表：每条记录一行
    long_table = []
    for year in years:
        data = data_by_year.get(year, {})
        for section, subject in sorted_subjects:
            v = data.get(section, {}).get(subject)
            if isinstance(v, list):
                v = v[0] if v else None
            if v is None:
                continue
            label, kind = SECTION_LABELS.get(section, (section, ""))
            long_table.append({
                "年份": year,
                "报表类型": label,
                "会计科目": subject,
                "类别": kind,
                "数值": v,
            })

    return wide_table, long_table


def export_subjects_to_csv(data_by_year, years, output_path):
    """导出长格式科目表为 CSV（适合导入数据库/BI 工具）"""
    import csv
    _, long_table = extract_subjects_table(data_by_year, years)
    if not long_table:
        return False
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["年份", "报表类型",
                                               "会计科目", "类别", "数值"])
        writer.writeheader()
        for row in long_table:
            writer.writerow(row)
    return True


# ============================================================
# 原始报表(万元)提取
# ============================================================

def extract_raw_table_wan(file_path, sheet_type="BS", divide_by=10000):
    """从原始 Excel 提取所有行（支持多 sheet）

    保留完整的原始科目结构（不标准化），包括：
    - 所有非标准科目（如"短期投资""预付款项""累计折旧"等）
    - 分类标题行（如"流动资产：""非流动资产：")
    - 空行

    支持多 sheet：自动遍历所有 sheet，合并属于指定 sheet_type 的内容。
    例如：BS 拆成"资产负债表"和"资产负债表（续）"两个 sheet 时也能完整提取。

    Args:
        file_path: Excel 文件路径
        sheet_type: 'BS'（资产负债表）或 'IS'（利润表）或 'CF'（现金流量表）
        divide_by: 数值除数，默认 10000（转万元），传 1 保持原始单位（元）

    Returns:
        rows: list of dict
            {
                'side': 'L' / 'R',  # 双栏布局：左/右侧；单栏：'L'
                'row_type': 'data' / 'header' / 'blank',
                'account': 科目名（可能为空）,
                'row_no': 行次（如果有）,
                'end_val': 期末值（divide_by 后）,
                'start_val': 年初值（divide_by 后，如果有）,
            }
    """
    rows = []

    try:
        xls = pd.ExcelFile(file_path, engine='openpyxl')
    except Exception:
        return rows

    # 遍历所有 sheet，提取属于指定 sheet_type 的内容
    for sheet_name in xls.sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name,
                               header=None, dtype=str)
        except Exception:
            continue

        # 判断是否属于目标 sheet_type
        detected = detect_sheet_type(sheet_name, df)
        if detected != sheet_type:
            continue

        # 判断这个 sheet 是 "资产/左侧" 还是 "负债+权益/右侧"
        # 逻辑：包含"（续）"、以"负债"或"所有者"开头、明确包含"负债部分" → 'R'
        # 其他（资产负债表、流动资产） → 'L'
        side = 'L'
        if sheet_type == "BS":
            # 续表表示（如"资产负债表（续）"）
            if "（续）" in sheet_name or "续表" in sheet_name:
                side = 'R'
            # 以"负债"或"所有者"开头
            elif sheet_name.startswith("负债") or sheet_name.startswith("所有者"):
                side = 'R'

        # 提取这个 sheet 的行（传 side 参数）
        sheet_rows = _extract_raw_rows_from_df(
            df, sheet_type, divide_by, side=side)
        rows.extend(sheet_rows)

    return rows


def _extract_raw_rows_from_df(df, sheet_type, divide_by, side='L'):
    """从单个 sheet 的 DataFrame 提取所有行（私有 helper）

    支持双栏和单栏布局，分别处理 BS/IS/CF。

    Args:
        df: pandas DataFrame
        sheet_type: 'BS'/'IS'/'CF'
        divide_by: 数值除数
        side: 'L' 或 'R'。单栏布局下，这个 sheet 的内容标记为哪一边。
             默认 'L'（资产或独立表）。
    """
    rows = []

    if sheet_type == "BS" and is_dual_column_layout(df):
        # 找列位置
        header_row_idx = None
        for i, row in df.iterrows():
            row_str = " ".join([str(v) for v in row.values
                                if v is not None and not (isinstance(v, float) and pd.isna(v))])
            if "资产" in row_str and "负债" in row_str and ("期末" in row_str or "余额" in row_str):
                header_row_idx = i
                break

        if header_row_idx is None:
            return rows

        header_row = df.iloc[header_row_idx]
        asset_col = None
        asset_end_col = None
        asset_start_col = None
        asset_row_col = None  # 行次
        liab_col = None
        liab_end_col = None
        liab_start_col = None
        liab_row_col = None

        for col_idx, val in header_row.items():
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            val_str = str(val).strip()
            if val_str == "资产" or val_str.startswith("资产"):
                if asset_col is None:
                    asset_col = col_idx
            elif "负债" in val_str and "所有者权益" in val_str:
                if liab_col is None and asset_col is not None:
                    liab_col = col_idx
            elif val_str == "行次":
                if asset_row_col is None:
                    asset_row_col = col_idx
                elif liab_row_col is None:
                    liab_row_col = col_idx
            elif "期末" in val_str:
                if asset_end_col is None:
                    asset_end_col = col_idx
                elif liab_end_col is None:
                    liab_end_col = col_idx
            elif "年初" in val_str:
                if asset_start_col is None:
                    asset_start_col = col_idx
                elif liab_start_col is None:
                    liab_start_col = col_idx

        if asset_col is None:
            asset_col = 0
        if liab_col is None:
            liab_col = 4
        if asset_end_col is None:
            asset_end_col = asset_col + 2
        if liab_end_col is None:
            liab_end_col = liab_col + 2
        if asset_start_col is None:
            asset_start_col = asset_end_col + 1
        if liab_start_col is None:
            liab_start_col = liab_end_col + 1

        # 提取数据行（从表头下一行开始）
        for i, row in df.iterrows():
            if i <= header_row_idx:
                continue

            def _extract_side(name_col, end_col, start_col, row_col, side):
                try:
                    name = row.iloc[name_col] if name_col < len(row) else None
                except (IndexError, AttributeError):
                    name = None
                try:
                    end_v = parse_value(row.iloc[end_col]) if end_col < len(row) else None
                except (IndexError, AttributeError):
                    end_v = None
                try:
                    start_v = parse_value(row.iloc[start_col]) if start_col < len(row) else None
                except (IndexError, AttributeError):
                    start_v = None
                try:
                    row_no = None
                    if row_col is not None and row_col < len(row):
                        rv = row.iloc[row_col]
                        if rv is not None:
                            try:
                                row_no = int(float(str(rv)))
                            except (ValueError, TypeError):
                                row_no = None
                except (IndexError, AttributeError):
                    row_no = None

                if not name or not isinstance(name, str):
                    return None

                # 识别行类型
                stripped = name.strip()
                if not stripped:
                    row_type = 'blank'
                elif stripped.endswith('：') or stripped.endswith(':'):
                    row_type = 'subheader'  # 分类标题
                elif any(kw in stripped for kw in ['合计', '总计']):
                    row_type = 'subtotal'
                else:
                    row_type = 'data'

                end_val = round(end_v / divide_by, 2) if end_v is not None else None
                start_val = round(start_v / divide_by, 2) if start_v is not None else None

                return {
                    'side': side,
                    'row_type': row_type,
                    'account': stripped,
                    'row_no': row_no,
                    'df_index': i,
                    'end_val': end_val,
                    'start_val': start_val,
                }

            left = _extract_side(asset_col, asset_end_col, asset_start_col, asset_row_col, 'L')
            if left:
                rows.append(left)
            right = _extract_side(liab_col, liab_end_col, liab_start_col, liab_row_col, 'R')
            if right:
                rows.append(right)
    else:
        # 单栏（IS / CF）
        # 找表头
        header_row_idx = None
        for i, row in df.iterrows():
            row_str = " ".join([str(v) for v in row.values
                                if v is not None and not (isinstance(v, float) and pd.isna(v))])
            # 去空格后再判断（避免"项 目"中间空格漏匹配）
            row_str_clean = row_str.replace(" ", "").replace("\u3000", "")
            if "项目" in row_str_clean or "科目" in row_str_clean or "本年" in row_str_clean:
                if any(kw in row_str_clean for kw in ["行次", "金额", "累计", "余额"]):
                    header_row_idx = i
                    break
        if header_row_idx is None:
            # 退回到第 4 行
            header_row_idx = 3

        header_row = df.iloc[header_row_idx]
        account_col = None
        row_col = None
        val_cols = []

        for col_idx, val in header_row.items():
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            val_str = str(val).strip()
            if val_str in ("项目", "科目") or "项目" in val_str:
                if account_col is None:
                    account_col = col_idx
            elif val_str == "行次":
                if row_col is None:
                    row_col = col_idx
            elif "金额" in val_str or "累计" in val_str:
                val_cols.append(col_idx)

        if account_col is None:
            account_col = 0
        if not val_cols:
            # 默认第 3、4 列
            val_cols = [account_col + 2, account_col + 3]

        for i, row in df.iterrows():
            if i <= header_row_idx:
                continue
            try:
                name = row.iloc[account_col] if account_col < len(row) else None
            except (IndexError, AttributeError):
                continue
            if not name or not isinstance(name, str):
                continue
            stripped = name.strip()
            if not stripped:
                continue

            try:
                row_no = None
                if row_col is not None and row_col < len(row):
                    rv = row.iloc[row_col]
                    if rv is not None:
                        try:
                            row_no = int(float(str(rv)))
                        except (ValueError, TypeError):
                            row_no = None
            except (IndexError, AttributeError):
                row_no = None

            vals = []
            for vc in val_cols:
                try:
                    v = parse_value(row.iloc[vc]) if vc < len(row) else None
                except (IndexError, AttributeError):
                    v = None
                vals.append(round(v / divide_by, 2) if v is not None else None)

            if any(kw in stripped for kw in ['合计', '总计']):
                row_type = 'subtotal'
            elif stripped.endswith('：'):
                row_type = 'subheader'
            else:
                row_type = 'data'

            # 统一字段名：双栏用 end_val/start_val，单栏也用同样的字段名
            end_val = vals[0] if len(vals) > 0 else None
            start_val = vals[1] if len(vals) > 1 else None

            rows.append({
                'side': side,
                'row_type': row_type,
                'account': stripped,
                'row_no': row_no,
                'df_index': i,  # 原始 df 物理行号，用于按报表顺序排序
                'end_val': end_val,
                'start_val': start_val,
                'values': vals,  # 保留以兼容
            })

    return rows


def extract_raw_table_wan_multi(file_info, years, sheet_type, divide_by=10000):
    """从多年文件提取原始报表(万元)，按行次合并

    Args:
        file_info: {year: {section: filepath}} 文件路径字典
        years: 排序后的年份列表
        sheet_type: 'BS' / 'IS' / 'CF'
        divide_by: 数值除数，10000 表示万元，1 表示元

    Returns:
        merged_rows: list of dict，每个 row 含
            'account', 'row_no', 'row_type', 'side',
            'values': {year: [end_val, start_val]} (BS) 或 {year: [val1, val2, ...]} (IS)
    """
    # 收集每个年份的文件
    year_data = {}
    for year in years:
        if year in file_info and sheet_type in file_info[year]:
            path = file_info[year][sheet_type]
            try:
                year_data[year] = extract_raw_table_wan(
                    path, sheet_type, divide_by=divide_by)
            except Exception:
                year_data[year] = []

    if not year_data:
        return []

    # 用最新年份作为模板（行结构）
    template_year = years[-1]
    if template_year not in year_data:
        return []
    template_rows = year_data[template_year]

    # 合并：按 (side, row_no) 匹配
    # 如果 row_no 是 None，用 account 匹配
    merged = []
    for tr in template_rows:
        key = (tr['side'], tr['row_no'], tr['account'])
        merged_row = {
            'side': tr['side'],
            'row_type': tr['row_type'],
            'account': tr['account'],
            'row_no': tr['row_no'],
            'values': {}
        }
        # 模板年份的数据
        if sheet_type == "BS":
            merged_row['values'][template_year] = {
                'end': tr.get('end_val'),
                'start': tr.get('start_val'),
            }
        else:
            merged_row['values'][template_year] = tr.get('values', [])

        # 匹配其他年份
        for other_year, other_rows in year_data.items():
            if other_year == template_year:
                continue
            match = None
            for or_ in other_rows:
                if or_['side'] == tr['side'] and or_['account'] == tr['account']:
                    if tr['row_no'] is not None and or_['row_no'] is not None and or_['row_no'] == tr['row_no']:
                        match = or_
                        break
                    elif tr['row_no'] is None and or_['row_no'] is None:
                        match = or_
                        break
            if match:
                if sheet_type == "BS":
                    merged_row['values'][other_year] = {
                        'end': match.get('end_val'),
                        'start': match.get('start_val'),
                    }
                else:
                    merged_row['values'][other_year] = match.get('values', [])

        merged.append(merged_row)

    return merged


# ============================================================
# 同比变化 Sheet 生成
# ============================================================

def _yl(y):
    """统一的年份字符串表示（处理 tuple/int）"""
    if isinstance(y, tuple):
        return f"{y[0]}-{str(y[1]).zfill(2)}"
    return str(y)


def _write_change_sheet(wb, data_by_year, years, file_info,
                          header_font, header_fill, cat_font, cat_fill,
                          divide_by=10000):
    """生成「期末较期初变化」Sheet

    对每份报表的“期末余额”与“年初余额”两列直接计算变化率和变化额。
    不需要多年数据（单份报表也能看期初→期末变化）。

    Args:
        file_info: {year: {section: filepath}}
    """
    if not file_info:
        ws = wb.create_sheet("期末较期初变化")
        ws.cell(row=1, column=1, value="⚠️ 无文件信息")
        ws.column_dimensions["A"].width = 50
        return

    # 收集每个 (year, section, account) → (期末, 期初)
    # 从原始 Excel 文件读期初/期末（不依赖 data_by_year）
    period_data = {}
    for year, sections in file_info.items():
        for section, path in sections.items():
            if section not in ('BS', 'IS', 'CF'):
                continue
            try:
                rows = extract_raw_table_wan(path, section, divide_by=1)
            except Exception:
                continue
            for r in rows:
                if r.get('row_type') in ('subheader', 'blank'):
                    continue
                account = r.get('account', '').strip()
                if not account:
                    continue
                # 提取期初/期末
                if section == 'BS':
                    end = r.get('end_val')
                    start = r.get('start_val')
                else:
                    vals = r.get('values', [])
                    end = vals[0] if len(vals) > 0 else None
                    start = vals[1] if len(vals) > 1 else None
                if end is None and start is None:
                    continue
                period_data[(year, section, account)] = (end, start)

    if not period_data:
        ws = wb.create_sheet("期末较期初变化")
        ws.cell(row=1, column=1, value="⚠️ 未能从报表提取期初/期末数据")
        ws.column_dimensions["A"].width = 50
        return

    # 收集所有原始 row（含 side, df_index, account）
    all_raw_rows = []
    for year, sections in file_info.items():
        for section, path in sections.items():
            if section not in ('BS', 'IS', 'CF'):
                continue
            try:
                rows = extract_raw_table_wan(path, section, divide_by=1)
            except Exception:
                continue
            for r in rows:
                if r.get('row_type') in ('subheader', 'blank'):
                    continue
                account = r.get('account', '').strip()
                if not account:
                    continue
                # 提取期初/期末
                if section == 'BS':
                    end = r.get('end_val')
                    start = r.get('start_val')
                else:
                    vals = r.get('values', [])
                    end = vals[0] if len(vals) > 0 else None
                    start = vals[1] if len(vals) > 1 else None
                if end is None and start is None:
                    continue
                all_raw_rows.append({
                    'year': year,
                    'section': section,
                    'side': r.get('side', 'L'),
                    'account': account,
                    'df_index': r.get('df_index', 0),
                    'end': end,
                    'start': start,
                })

    if not all_raw_rows:
        ws = wb.create_sheet("期末较期初变化")
        ws.cell(row=1, column=1, value="⚠️ 未能从报表提取期初/期末数据")
        ws.column_dimensions["A"].width = 50
        return

    # 按报表物理顺序排序：
    # - BS: 先 side='L'（资产）按 df_index，再 side='R'（负债）按 df_index
    # - IS/CF: 只有一个 side='L'，按 df_index
    section_order = {'BS': 0, 'IS': 1, 'CF': 2}
    side_order = {'L': 0, 'R': 1}
    sorted_rows = sorted(
        all_raw_rows,
        key=lambda x: (section_order.get(x['section'], 99),
                        side_order.get(x['side'], 0),
                        x['df_index']))

    unit_label = "万元" if divide_by == 10000 else "元"

    ws = wb.create_sheet("期末较期初变化")
    # 顶部加目录
    toc_parts = []
    for st, sl in [("BS", "资产负债表"), ("IS", "利润表"), ("CF", "现金流量表")]:
        year_list = [y for y in years if y in file_info and st in file_info[y]]
        if year_list:
            year_str = ", ".join(_yl(y) for y in year_list)
            toc_parts.append(f"{sl} ({year_str})")
    if toc_parts:
        toc_cell = ws.cell(row=1, column=1, value="📋 目录: " + " | ".join(toc_parts))
        toc_cell.font = Font(bold=True, color="FFFFFF", size=11)
        toc_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws.row_dimensions[1].height = 22
    # 表头
    headers = ["报表类型", "会计科目", "单位", "年份",
               "期初余额", "期末余额", "变化额", "变化率"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 数据（从第 3 行开始，第 1 行是目录，第 2 行是表头）
    row = 3
    prev_section = None
    prev_side = None
    for r in sorted_rows:
        section = r['section']
        side = r['side']
        if section != prev_section:
            if prev_section is not None:
                row += 1
            # 加分页提示（醒目大字）
            if prev_section is not None:  # 不是第一个分类才加分页
                section_emoji = "💰" if section == "IS" else "💵"
                section_color = "548235" if section == "IS" else "7030A0"
                divider_row = row
                cell = ws.cell(row=divider_row, column=1,
                                value=f"{section_emoji} {section_emoji} {section_emoji} 【{SECTION_LABELS[section][0]}】 {section_emoji} {section_emoji} {section_emoji}")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color=section_color, end_color=section_color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(
                    start_row=divider_row, start_column=1,
                    end_row=divider_row, end_column=len(headers))
                ws.row_dimensions[divider_row].height = 25
                row += 1
            cell = ws.cell(row=row, column=1, value=SECTION_LABELS[section][0])
            cell.font = cat_font
            cell.fill = cat_fill
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row,
                end_column=len(headers))
            row += 1
            prev_section = section
            prev_side = None
        end = r['end']
        start = r['start']
        ws.cell(row=row, column=1, value=SECTION_LABELS[section][0])
        ws.cell(row=row, column=2, value=r['account'])
        ws.cell(row=row, column=3, value=unit_label)
        ws.cell(row=row, column=4, value=_yl(r['year']))
        # 期初
        cell = ws.cell(row=row, column=5)
        if start is not None:
            cell.value = float(start / divide_by)
            cell.number_format = "#,##0.00"
        else:
            cell.value = "-"
        # 期末
        cell = ws.cell(row=row, column=6)
        if end is not None:
            cell.value = float(end / divide_by)
            cell.number_format = "#,##0.00"
        else:
            cell.value = "-"
        # 变化额
        cell = ws.cell(row=row, column=7)
        if start is not None and end is not None:
            cell.value = float((end - start) / divide_by)
            cell.number_format = "#,##0.00;[Red]-#,##0.00"
        else:
            cell.value = "-"
        # 变化率
        cell = ws.cell(row=row, column=8)
        if start is not None and end is not None and start != 0:
            change = (end - start) / abs(start) * 100
            if change > 0.5:
                cell.value = f"↑ {change:+.2f}%"
            elif change < -0.5:
                cell.value = f"↓ {change:+.2f}%"
            else:
                cell.value = f"→ {change:+.2f}%"
        else:
            cell.value = "-"
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14


# ============================================================
# 报告生成
# ============================================================
def _header_style():
    return {
        "header_font": Font(bold=True, size=12, color="FFFFFF"),
        "header_fill": PatternFill(start_color="1F4E78",
                                   end_color="1F4E78", fill_type="solid"),
        "cat_fill": PatternFill(start_color="2E75B6",
                                end_color="2E75B6", fill_type="solid"),
        "cat_font": Font(bold=True, color="FFFFFF"),
        "up_fill": PatternFill(start_color="C6EFCE",
                               end_color="C6EFCE", fill_type="solid"),
        "down_fill": PatternFill(start_color="FFC7CE",
                                 end_color="FFC7CE", fill_type="solid"),
    }


def generate_report(data_by_year, indicators_by_year, years, trends,
                    output_path, company_name="目标公司",
                    analyst_name="", errors=None):
    """生成 Excel 报告（多年 + 趋势 + 图表）

    Args:
        analyst_name: 分析人/作者名，会显示在首页
    """
    wb = Workbook()
    style = _header_style()
    header_font = style["header_font"]
    header_fill = style["header_fill"]
    cat_fill = style["cat_fill"]
    cat_font = style["cat_font"]
    up_fill = style["up_fill"]

    def _yl(y):  # noqa: F841 - 保留以向后兼容
        """统一的年份字符串表示（处理 tuple/int）"""
        if isinstance(y, tuple):
            return f"{y[0]}-{str(y[1]).zfill(2)}"
        return str(y)
    down_fill = style["down_fill"]

    # ===== 首页 =====
    ws0 = wb.active
    ws0.title = "首页"
    ws0["A1"] = f"{company_name} - 财务分析报告"
    ws0["A1"].font = Font(bold=True, size=18, color="1F4E78")
    ws0.merge_cells("A1:F1")
    ws0["A2"] = f"分析年份: {' / '.join(map(str, years))}  |  生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws0["A2"].font = Font(italic=True, color="888888")
    ws0.merge_cells("A2:F2")
    # 工具版本号
    ws0["A3"] = f"🔧 工具版本: v{__version__}  |  报告生成日期: {datetime.now().strftime('%Y-%m-%d')}"
    ws0["A3"].font = Font(italic=True, color="666666", size=9)
    ws0.merge_cells("A3:F3")
    if analyst_name and analyst_name.strip():
        ws0["A4"] = f"👤 分析人: {analyst_name.strip()}"
        ws0["A4"].font = Font(bold=True, size=11, color="1F4E78")
        ws0.merge_cells("A4:F4")
        desc_start_row = 6
    else:
        desc_start_row = 5

    ws0.cell(row=desc_start_row, column=1, value="⚠️ 数据说明").font = Font(
        bold=True, size=12)
    notes = [
        "本报告由本地程序自动生成，原始数据来自您提供的财报文件。",
        "本工具不联网、不上传数据，所有计算均在您的电脑上完成。",
        "财务指标的取数逻辑详见各 Sheet 注释，建议结合附注和审计意见综合判断。",
        "趋势分析包含同比变化（YoY）和复合增长率（CAGR）。",
    ]
    for i, note in enumerate(notes, desc_start_row + 1):
        ws0.cell(row=i, column=1, value=f"  • {note}")

    struct_start = desc_start_row + len(notes) + 2
    ws0.cell(row=struct_start, column=1, value="📊 报告结构").font = Font(
        bold=True, size=12)
    structure = [
        ("原始数据", "三大报表原始数据（按年并列）"),
        ("财务指标", "盈利/偿债/营运/现金流/杜邦分析（按年并列）"),
        ("趋势分析", "各指标 YoY 同比 + CAGR 复合增长率"),
        ("趋势图表", "openpyxl 原生图表：营收/利润/利润率趋势"),
        ("结构图表", "资产/负债/利润结构图"),
        ("数据校验", "资产负债表平衡校验、关键指标健康度"),
    ]
    for i, (sheet, desc) in enumerate(structure, struct_start + 2):
        ws0.cell(row=i, column=1, value=sheet).font = Font(bold=True, color="1F4E78")
        ws0.cell(row=i, column=2, value=desc)
    ws0.column_dimensions["A"].width = 20
    ws0.column_dimensions["B"].width = 60

    # ===== 原始数据（元单位，保留所有原始行）=====
    # 从 data_by_year 中提取 __files__ 信息
    file_info = {}
    for y in years:
        yd = data_by_year.get(y, {})
        if isinstance(yd, dict) and '__files__' in yd:
            file_info[y] = yd['__files__']

    # 原始数据 Sheet（元单位，保留所有原始行）
    ws1 = wb.create_sheet("原始数据")
    # 顶部加目录行（点哪个跳转哪个区域）
    toc_parts = []
    if file_info:
        for st, sl in [("BS", "资产负债表"), ("IS", "利润表"), ("CF", "现金流量表")]:
            year_list = [y for y in years if y in file_info and st in file_info[y]]
            if year_list:
                year_str = ", ".join(_yl(y) for y in year_list)
                toc_parts.append(f"{sl} ({year_str})")
    if toc_parts:
        toc_cell = ws1.cell(row=1, column=1, value="📋 目录: " + " | ".join(toc_parts))
        toc_cell.font = Font(bold=True, color="FFFFFF", size=11)
        toc_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        ws1.row_dimensions[1].height = 22
    toc_offset = 2  # 后续数据从第 2 行开始

    for sheet_type, sheet_label in [("BS", "资产负债表"), ("IS", "利润表"),
                                      ("CF", "现金流量表")]:
        merged_rows = extract_raw_table_wan_multi(
            file_info, years, sheet_type, divide_by=1)  # 1 表示保持元
        if not merged_rows:
            continue

        if sheet_type == "BS":
            # 双栏：左侧资产、右侧负债+权益
            # 表头从 row 2 开始（row 1 是目录）
            headers = ["资产(项目)", "期末余额(元)", "年初余额(元)",
                       f"{sheet_label}(项目)", "期末余额(元)", "年初余额(元)"]
            for col, h in enumerate(headers, 1):
                cell = ws1.cell(row=2, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            left_rows = [r for r in merged_rows if r['side'] == 'L']
            right_rows = [r for r in merged_rows if r['side'] == 'R']
            max_rows = max(len(left_rows), len(right_rows))
            for i in range(max_rows):
                row = i + 3
                if i < len(left_rows):
                    lr = left_rows[i]
                    cell = ws1.cell(row=row, column=1, value=lr['account'])
                    if lr['row_type'] == 'subheader':
                        cell.font = cat_font
                        cell.fill = cat_fill
                        ws1.merge_cells(
                            start_row=row, start_column=1, end_row=row,
                            end_column=3)
                        continue
                    if lr['row_type'] == 'subtotal':
                        cell.font = Font(bold=True)
                    if lr['values']:
                        latest_y = years[-1]
                        v = lr['values'].get(latest_y, {}).get('end')
                        cell = ws1.cell(row=row, column=2)
                        if v is None:
                            cell.value = "-"
                        else:
                            cell.value = float(v)
                            cell.number_format = "#,##0.00"
                        v2 = lr['values'].get(latest_y, {}).get('start')
                        cell = ws1.cell(row=row, column=3)
                        if v2 is None:
                            cell.value = "-"
                        else:
                            cell.value = float(v2)
                            cell.number_format = "#,##0.00"
                if i < len(right_rows):
                    rr = right_rows[i]
                    cell = ws1.cell(row=row, column=4, value=rr['account'])
                    if rr['row_type'] == 'subheader':
                        cell.font = cat_font
                        cell.fill = cat_fill
                        ws1.merge_cells(
                            start_row=row, start_column=4, end_row=row,
                            end_column=6)
                        continue
                    if rr['row_type'] == 'subtotal':
                        cell.font = Font(bold=True)
                    if rr['values']:
                        latest_y = years[-1]
                        v = rr['values'].get(latest_y, {}).get('end')
                        cell = ws1.cell(row=row, column=5)
                        if v is None:
                            cell.value = "-"
                        else:
                            cell.value = float(v)
                            cell.number_format = "#,##0.00"
                        v2 = rr['values'].get(latest_y, {}).get('start')
                        cell = ws1.cell(row=row, column=6)
                        if v2 is None:
                            cell.value = "-"
                        else:
                            cell.value = float(v2)
                            cell.number_format = "#,##0.00"
            ws1.column_dimensions["A"].width = 30
            ws1.column_dimensions["B"].width = 16
            ws1.column_dimensions["C"].width = 16
            ws1.column_dimensions["D"].width = 30
            ws1.column_dimensions["E"].width = 16
            ws1.column_dimensions["F"].width = 16
            # 之后是 IS/CF 块，需要追加到同一个 Sheet
        else:
            # 单栏：IS / CF - 追加到"原始数据" Sheet
            # 加醒目的分页提示（3 行高度 + 大字背景）
            last_row = ws1.max_row
            # 空行
            for r in range(last_row + 1, last_row + 4):
                ws1.cell(row=r, column=1, value="")
            # 大字分页提示
            divider_row = last_row + 4
            section_color = "548235" if sheet_type == "IS" else "7030A0"
            section_emoji = "💰" if sheet_type == "IS" else "💵"
            cell = ws1.cell(row=divider_row, column=1,
                            value=f"{section_emoji} {section_emoji} {section_emoji} 下面开始是【{sheet_label}】数据 {section_emoji} {section_emoji} {section_emoji}")
            cell.font = Font(bold=True, color="FFFFFF", size=13)
            cell.fill = PatternFill(start_color=section_color, end_color=section_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws1.merge_cells(start_row=divider_row, start_column=1,
                            end_row=divider_row, end_column=len(years) + 1)
            ws1.row_dimensions[divider_row].height = 30

            # 块标题
            start_row = last_row + 6
            cell = ws1.cell(row=start_row, column=1,
                            value=f"【{sheet_label}】")
            cell.font = cat_font
            cell.fill = cat_fill
            ws1.merge_cells(start_row=start_row, start_column=1,
                            end_row=start_row, end_column=len(years) + 1)

            # 表头
            header_row = start_row + 1
            ws1.cell(row=header_row, column=1, value="项目")
            for col_idx, y in enumerate(years, 2):
                ws1.cell(row=header_row, column=col_idx,
                         value=f"{_yl(y)}(元)")
            for col_idx in range(1, len(years) + 2):
                c = ws1.cell(row=header_row, column=col_idx)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")

            # 数据
            cur_row = header_row + 1
            for r in merged_rows:
                cell = ws1.cell(row=cur_row, column=1, value=r['account'])
                if r['row_type'] == 'subheader':
                    cell.font = cat_font
                    cell.fill = cat_fill
                    ws1.merge_cells(
                        start_row=cur_row, start_column=1, end_row=cur_row,
                        end_column=len(years) + 1)
                    cur_row += 1
                    continue
                if r['row_type'] == 'subtotal':
                    cell.font = Font(bold=True)
                for col_idx, y in enumerate(years, 2):
                    vals = r['values'].get(y, [])
                    v = vals[0] if vals else None
                    cell = ws1.cell(row=cur_row, column=col_idx)
                    if v is None:
                        cell.value = "-"
                    else:
                        cell.value = float(v)
                        cell.number_format = "#,##0.00"
                cur_row += 1

    # ===== 原始报表(万元) - 保留所有原始科目行 =====
    # 从 data_by_year 中提取 __files__ 信息
    file_info = {}
    for y in years:
        yd = data_by_year.get(y, {})
        if isinstance(yd, dict) and '__files__' in yd:
            file_info[y] = yd['__files__']

    for sheet_type, sheet_label in [("BS", "资产负债表"), ("IS", "利润表"),
                                      ("CF", "现金流量表")]:
        merged_rows = extract_raw_table_wan_multi(file_info, years, sheet_type)
        if not merged_rows:
            continue

        sheet_name = f"原始报表(万元)-{sheet_label}"
        ws_raw = wb.create_sheet(sheet_name)

        if sheet_type == "BS":
            headers = ["资产(项目)", "期末余额", "年初余额",
                       f"{sheet_label}(项目)", "期末余额", "年初余额"]
            for col, h in enumerate(headers, 1):
                cell = ws_raw.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            left_rows = [r for r in merged_rows if r['side'] == 'L']
            right_rows = [r for r in merged_rows if r['side'] == 'R']

            max_rows = max(len(left_rows), len(right_rows))
            for i in range(max_rows):
                row = i + 2
                if i < len(left_rows):
                    lr = left_rows[i]
                    cell = ws_raw.cell(row=row, column=1, value=lr['account'])
                    if lr['row_type'] == 'subheader':
                        cell.font = cat_font
                        cell.fill = cat_fill
                        ws_raw.merge_cells(
                            start_row=row, start_column=1, end_row=row,
                            end_column=3)
                        continue
                    if lr['row_type'] == 'subtotal':
                        cell.font = Font(bold=True)
                    if lr['values']:
                        latest_y = years[-1]
                        v = lr['values'].get(latest_y, {}).get('end')
                        cell = ws_raw.cell(row=row, column=2)
                        if v is None:
                            cell.value = "-"
                        else:
                            cell.value = float(v)
                            cell.number_format = "#,##0.00"
                        v2 = lr['values'].get(latest_y, {}).get('start')
                        cell = ws_raw.cell(row=row, column=3)
                        if v2 is None:
                            cell.value = "-"
                        else:
                            cell.value = float(v2)
                            cell.number_format = "#,##0.00"
                if i < len(right_rows):
                    rr = right_rows[i]
                    cell = ws_raw.cell(row=row, column=4, value=rr['account'])
                    if rr['row_type'] == 'subheader':
                        cell.font = cat_font
                        cell.fill = cat_fill
                        ws_raw.merge_cells(
                            start_row=row, start_column=4, end_row=row,
                            end_column=6)
                        continue
                    if rr['row_type'] == 'subtotal':
                        cell.font = Font(bold=True)
                    if rr['values']:
                        latest_y = years[-1]
                        v = rr['values'].get(latest_y, {}).get('end')
                        cell = ws_raw.cell(row=row, column=5)
                        if v is None:
                            cell.value = "-"
                        else:
                            cell.value = float(v)
                            cell.number_format = "#,##0.00"
                        v2 = rr['values'].get(latest_y, {}).get('start')
                        cell = ws_raw.cell(row=row, column=6)
                        if v2 is None:
                            cell.value = "-"
                        else:
                            cell.value = float(v2)
                            cell.number_format = "#,##0.00"

            ws_raw.column_dimensions["A"].width = 30
            ws_raw.column_dimensions["B"].width = 14
            ws_raw.column_dimensions["C"].width = 14
            ws_raw.column_dimensions["D"].width = 30
            ws_raw.column_dimensions["E"].width = 14
            ws_raw.column_dimensions["F"].width = 14
        else:
            headers = ["项目"]
            for y in years:
                yl = _yl(y)
                headers.append(f"{yl}(万元)")
            for col, h in enumerate(headers, 1):
                cell = ws_raw.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for i, r in enumerate(merged_rows, 2):
                cell = ws_raw.cell(row=i, column=1, value=r['account'])
                if r['row_type'] == 'subheader':
                    cell.font = cat_font
                    cell.fill = cat_fill
                    ws_raw.merge_cells(
                        start_row=i, start_column=1, end_row=i,
                        end_column=len(years) + 1)
                    continue
                if r['row_type'] == 'subtotal':
                    cell.font = Font(bold=True)
                for col_idx, y in enumerate(years, 2):
                    vals = r['values'].get(y, [])
                    v = vals[0] if vals else None
                    cell = ws_raw.cell(row=i, column=col_idx)
                    if v is None:
                        cell.value = "-"
                    else:
                        cell.value = float(v)
                        cell.number_format = "#,##0.00"
            ws_raw.column_dimensions["A"].width = 35
            for col_idx in range(2, len(years) + 2):
                ws_raw.column_dimensions[get_column_letter(col_idx)].width = 16

    # ===== 财务指标（按年并列）=====
    ws2 = wb.create_sheet("财务指标")
    headers = ["指标", "维度"] + [str(y) for y in years]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    categories = {
        "📈 盈利能力": [
            "毛利率", "净利率", "营业利润率", "销售费用率", "管理费用率",
            "研发费用率", "财务费用率", "ROE (归母净资产收益率)",
            "ROA (总资产收益率)", "ROIC (投入资本回报率, 估算)", "归母净利润/营业收入",
            "EBITDA 利润率",
        ],
        "💰 偿债能力": [
            "流动比率", "速动比率", "现金比率", "资产负债率",
            "权益乘数", "长期资本适合率", "利息保障倍数", "商誉/净资产",
        ],
        "⚙️ 营运能力": [
            "应收账款周转率(次)", "应收账款周转天数(天)",
            "存货周转率(次)", "存货周转天数(天)",
            "总资产周转率(次)", "总资产周转天数(天)",
            "固定资产周转率(次)", "流动资产周转率(次)", "应收账款/营收",
        ],
        "💵 现金流": [
            "经营现金流/净利润(现金含量)", "自由现金流(FCF)",
            "经营现金流/营业收入", "资本支出/营业收入", "经营现金流/总资产",
        ],
        "🚀 发展能力": [
            "营业总收入增长率", "净利润增长率", "归母净利润增长率",
        ],
        "🔍 杜邦分析": [
            "杜邦 ROE(%)", "杜邦分解-净利率(%)",
            "杜邦分解-总资产周转率", "杜邦分解-权益乘数",
        ],
    }

    unit_map = {
        "%": ["毛利率", "净利率", "营业利润率", "销售费用率", "管理费用率",
              "研发费用率", "财务费用率", "ROE", "ROA", "ROIC",
              "归母净利润/营业收入", "资产负债率",
              "经营现金流/营业收入", "资本支出/营业收入", "经营现金流/总资产",
              "杜邦 ROE(%)", "杜邦分解-净利率(%)",
              "EBITDA 利润率", "商誉/净资产", "应收账款/营收",
              "营业总收入增长率", "净利润增长率", "归母净利润增长率"],
        "倍": ["利息保障倍数"],
        "天": ["应收账款周转天数(天)", "存货周转天数(天)", "总资产周转天数(天)"],
        "次": ["应收账款周转率(次)", "存货周转率(次)", "总资产周转率(次)",
               "固定资产周转率(次)", "流动资产周转率(次)"],
        "元": ["自由现金流(FCF)", "毛利"],
    }

    def get_unit(name):
        for unit, keywords in unit_map.items():
            for kw in keywords:
                if kw in name:
                    return unit
        return ""

    row = 2
    for cat, items in categories.items():
        cell = ws2.cell(row=row, column=1, value=cat)
        cell.font = cat_font
        cell.fill = cat_fill
        ws2.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=len(years) + 2)
        row += 1
        for item in items:
            ws2.cell(row=row, column=1, value=item)
            ws2.cell(row=row, column=2, value=get_unit(item))
            for col_idx, year in enumerate(years, 3):
                v = indicators_by_year.get(year, {}).get(item)
                cell = ws2.cell(row=row, column=col_idx)
                if v is None or v == 0:
                    cell.value = "-"
                elif isinstance(v, str):
                    cell.value = v
                else:
                    cell.value = v
                    unit = get_unit(item)
                    if unit == "%":
                        cell.number_format = "0.00"
                    elif unit == "天":
                        cell.number_format = "0.0"
                    else:
                        cell.number_format = "#,##0.00"
            row += 1
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 8
    for col_idx in range(3, len(years) + 3):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 15

    # ===== 趋势分析 =====
    if trends:
        ws3 = wb.create_sheet("趋势分析")
        yoy_headers = [f"{years[i]}-{years[i+1]} YoY" for i in range(len(years) - 1)]
        headers = ["指标"] + [str(y) for y in years] + yoy_headers + ["CAGR"]
        for col, h in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row = 2
        # 按类别分组
        for cat, items in categories.items():
            cell = ws3.cell(row=row, column=1, value=cat)
            cell.font = cat_font
            cell.fill = cat_fill
            ws3.merge_cells(start_row=row, start_column=1,
                            end_row=row, end_column=len(headers))
            row += 1
            for item in items:
                if item not in trends:
                    continue
                t = trends[item]
                ws3.cell(row=row, column=1, value=item)
                # 各年数值
                for col_idx, (year, v) in enumerate(t["values"], 2):
                    cell = ws3.cell(row=row, column=col_idx)
                    if v is None:
                        cell.value = "-"
                    else:
                        cell.value = v
                        unit = get_unit(item)
                        if unit == "%":
                            cell.number_format = "0.00"
                        elif unit == "天":
                            cell.number_format = "0.0"
                        else:
                            cell.number_format = "#,##0.00"
                # YoY
                for col_idx, yoy_info in enumerate(t["yoy"]):
                    cell = ws3.cell(row=row, column=len(years) + 2 + col_idx)
                    if yoy_info["yoy"] is not None:
                        cell.value = f"{yoy_info['direction']} {yoy_info['yoy']:.2f}%"
                        if yoy_info["yoy"] > 0:
                            cell.fill = up_fill
                        elif yoy_info["yoy"] < 0:
                            cell.fill = down_fill
                # CAGR
                col_idx = len(headers)
                cell = ws3.cell(row=row, column=col_idx)
                if t["cagr"] is not None:
                    cell.value = f"{t['cagr']:.2f}%"
                    if t["cagr"] > 0:
                        cell.fill = up_fill
                    elif t["cagr"] < 0:
                        cell.fill = down_fill
                row += 1
        ws3.column_dimensions["A"].width = 35
        for col_idx in range(2, len(headers) + 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 15

    # ===== 趋势图表 =====
    if years and len(years) >= 1:
        ws_chart = wb.create_sheet("趋势图表")

        # 准备图表数据
        chart_data = {
            "营收 & 利润": {
                "营业收入": ("IS", "营业收入", "元"),
                "净利润": ("IS", "净利润", "元"),
                "归母净利润": ("IS", "归母净利润", "元"),
            },
            "利润率": {
                "毛利率": ("指标", "毛利率", "%"),
                "净利率": ("指标", "净利率", "%"),
                "营业利润率": ("指标", "营业利润率", "%"),
                "ROE": ("指标", "ROE (归母净资产收益率)", "%"),
            },
        }

        chart_row = 1
        for chart_title, series_def in chart_data.items():
            # 写表头
            ws_chart.cell(row=chart_row, column=1, value=chart_title).font = Font(
                bold=True, size=14, color="1F4E78")
            chart_row += 1
            ws_chart.cell(row=chart_row, column=1, value="指标")
            for col_idx, year in enumerate(years, 2):
                ws_chart.cell(row=chart_row, column=col_idx, value=str(year))
            chart_row += 1

            data_start_row = chart_row
            series_rows = []
            for series_name, (source_type, key, unit) in series_def.items():
                ws_chart.cell(row=chart_row, column=1, value=series_name)
                for col_idx, year in enumerate(years, 2):
                    if source_type == "IS":
                        v = data_by_year.get(year, {}).get("IS", {}).get(key)
                    else:
                        v = indicators_by_year.get(year, {}).get(key)
                    cell = ws_chart.cell(row=chart_row, column=col_idx)
                    if v is None:
                        cell.value = 0
                    else:
                        cell.value = v
                        if unit == "%":
                            cell.number_format = "0.00"
                        else:
                            cell.number_format = "#,##0"
                series_rows.append(chart_row)
                chart_row += 1
            data_end_row = chart_row - 1
            header_row = data_start_row - 1

            # 创建折线图
            chart = LineChart()
            chart.title = chart_title
            chart.style = 12
            chart.y_axis.title = unit
            chart.x_axis.title = "年份"
            chart.height = 10
            chart.width = 20

            for sr in series_rows:
                data_ref = Reference(ws_chart, min_col=2, max_col=len(years) + 1,
                                     min_row=sr, max_row=sr)
                chart.add_data(data_ref, titles_from_data=False)
                # 设置系列名称
                if chart.series:
                    chart.series[-1].tx = openpyxl.chart.series.SeriesLabel(
                        v=ws_chart.cell(row=sr, column=1).value
                    )

            cats = Reference(ws_chart, min_col=2, max_col=len(years) + 1,
                             min_row=header_row, max_row=header_row)
            chart.set_categories(cats)

            # 图表位置
            ws_chart.add_chart(chart, f"A{chart_row + 1}")
            chart_row += 22  # 给图表留位置

    # ===== 结构图表（最新一年）=====
    if years:
        latest_year = years[-1]
        latest_data = data_by_year.get(latest_year, {})

        ws_pie = wb.create_sheet("结构图表")

        # 资产结构
        ws_pie.cell(row=1, column=1,
                    value=f"📊 资产结构 ({latest_year})").font = Font(
            bold=True, size=14, color="1F4E78")

        bs = latest_data.get("BS", {})
        current_assets = get_value(bs, "流动资产合计")
        non_current_assets = get_value(bs, "非流动资产合计")
        total_assets = get_value(bs, "资产总计")

        if total_assets:
            ws_pie.cell(row=3, column=1, value="类别")
            ws_pie.cell(row=3, column=2, value="金额")
            ws_pie.cell(row=4, column=1, value="流动资产")
            ws_pie.cell(row=4, column=2, value=current_assets or 0)
            ws_pie.cell(row=5, column=1, value="非流动资产")
            ws_pie.cell(row=5, column=2, value=non_current_assets or 0)

            pie1 = PieChart()
            pie1.title = f"资产结构 {latest_year}"
            labels = Reference(ws_pie, min_col=1, min_row=4, max_row=5)
            data = Reference(ws_pie, min_col=2, min_row=4, max_row=5)
            pie1.add_data(data, titles_from_data=False)
            pie1.set_categories(labels)
            pie1.dataLabels = DataLabelList(showPercent=True)
            pie1.height = 8
            pie1.width = 12
            ws_pie.add_chart(pie1, "D3")

        # 负债 vs 权益
        total_liab = get_value(bs, "负债合计")
        total_equity = get_value(bs, "所有者权益合计")
        if total_liab and total_equity:
            ws_pie.cell(row=20, column=1,
                        value=f"📊 资本结构 ({latest_year})").font = Font(
                bold=True, size=14, color="1F4E78")
            ws_pie.cell(row=22, column=1, value="类别")
            ws_pie.cell(row=22, column=2, value="金额")
            ws_pie.cell(row=23, column=1, value="负债")
            ws_pie.cell(row=23, column=2, value=total_liab)
            ws_pie.cell(row=24, column=1, value="所有者权益")
            ws_pie.cell(row=24, column=2, value=total_equity)

            pie2 = PieChart()
            pie2.title = f"资本结构 {latest_year}"
            labels = Reference(ws_pie, min_col=1, min_row=23, max_row=24)
            data = Reference(ws_pie, min_col=2, min_row=23, max_row=24)
            pie2.add_data(data, titles_from_data=False)
            pie2.set_categories(labels)
            pie2.dataLabels = DataLabelList(showPercent=True)
            pie2.height = 8
            pie2.width = 12
            ws_pie.add_chart(pie2, "D22")

        # 利润结构
        is_ = latest_data.get("IS", {})
        revenue = get_value(is_, "营业收入")
        cost = get_value(is_, "营业成本")
        operating_profit = get_value(is_, "营业利润")
        net_profit = get_value(is_, "净利润")
        if revenue:
            ws_pie.cell(row=40, column=1,
                        value=f"📊 利润结构 ({latest_year})").font = Font(
                bold=True, size=14, color="1F4E78")
            ws_pie.cell(row=42, column=1, value="类别")
            ws_pie.cell(row=42, column=2, value="金额")
            ws_pie.cell(row=43, column=1, value="营业成本")
            ws_pie.cell(row=43, column=2, value=cost or 0)
            ws_pie.cell(row=44, column=1, value="营业利润")
            ws_pie.cell(row=44, column=2, value=operating_profit or 0)
            ws_pie.cell(row=45, column=1, value="净利润")
            ws_pie.cell(row=45, column=2, value=net_profit or 0)

            pie3 = PieChart()
            pie3.title = f"利润结构 {latest_year}"
            labels = Reference(ws_pie, min_col=1, min_row=43, max_row=45)
            data = Reference(ws_pie, min_col=2, min_row=43, max_row=45)
            pie3.add_data(data, titles_from_data=False)
            pie3.set_categories(labels)
            pie3.dataLabels = DataLabelList(showPercent=True)
            pie3.height = 8
            pie3.width = 12
            ws_pie.add_chart(pie3, "D42")

        ws_pie.column_dimensions["A"].width = 18
        ws_pie.column_dimensions["B"].width = 18

    # ===== 数据校验 =====
    ws4 = wb.create_sheet("数据校验")
    ws4.cell(row=1, column=1, value="数据健康度检查").font = Font(bold=True, size=14)
    ws4.merge_cells("A1:C1")
    for col, h in enumerate(["检查项", "结果", "状态"], 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row = 4
    for year in years:
        data = data_by_year.get(year, {})
        bs = data.get("BS", {})
        is_ = data.get("IS", {})
        cf = data.get("CF", {})

        ws4.cell(row=row, column=1,
                 value=f"═══ {year} 年 ═══").font = Font(bold=True, color="1F4E78")
        ws4.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=3)
        row += 1

        # 资产 = 负债 + 权益
        assets_total = get_value(bs, "资产总计")
        balance = get_value(bs, "负债和所有者权益总计")
        if assets_total and balance:
            diff = abs(assets_total - balance)
            ratio = diff / assets_total * 100
            status = "✅ 通过" if ratio < 0.1 else "⚠️ 异常"
            ws4.cell(row=row, column=1, value="资产=负债+权益")
            ws4.cell(row=row, column=2,
                     value=f"差异 {diff:.2f} ({ratio:.4f}%)")
            ws4.cell(row=row, column=3, value=status)
            row += 1

        # 净利率健康
        revenue = get_value(is_, "营业收入")
        net_profit = get_value(is_, "净利润")
        if revenue and net_profit:
            net_margin = net_profit / revenue * 100
            status = "✅ 正常" if -50 < net_margin < 50 else "⚠️ 异常"
            ws4.cell(row=row, column=1, value="净利率")
            ws4.cell(row=row, column=2, value=f"{net_margin:.2f}%")
            ws4.cell(row=row, column=3, value=status)
            row += 1

        # 现金含量
        op_cf = get_value(cf, "经营活动产生的现金流量净额")
        if op_cf and net_profit:
            ratio = op_cf / net_profit
            status = "✅ 良好" if ratio > 0.5 else "⚠️ 关注"
            ws4.cell(row=row, column=1, value="经营现金流/净利润")
            ws4.cell(row=row, column=2, value=f"{ratio:.2f}")
            ws4.cell(row=row, column=3, value=status)
            row += 1

        # 资产负债率健康
        total_liab = get_value(bs, "负债合计")
        if total_liab and assets_total:
            ratio = total_liab / assets_total * 100
            status = "✅ 正常" if ratio < 80 else "⚠️ 偏高"
            ws4.cell(row=row, column=1, value="资产负债率")
            ws4.cell(row=row, column=2, value=f"{ratio:.2f}%")
            ws4.cell(row=row, column=3, value=status)
            row += 1

        row += 1

    if errors:
        ws4.cell(row=row, column=1, value="═══ 解析错误 ═══").font = Font(
            bold=True, color="C00000")
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for filepath, err in errors:
            ws4.cell(row=row, column=1, value=Path(filepath).name)
            ws4.cell(row=row, column=2, value=err[:200])
            ws4.cell(row=row, column=3, value="❌")
            row += 1

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 40
    ws4.column_dimensions["C"].width = 15

    # ===== 科目表（宽格式）=====
    wide_table, long_table = extract_subjects_table(data_by_year, years)
    if wide_table:
        ws_wide = wb.create_sheet("科目表(宽)")
        # 表头
        headers = ["报表类型", "会计科目"] + [str(y) for y in years]
        for col, h in enumerate(headers, 1):
            cell = ws_wide.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # 数据 + 分类标题
        row = 2
        prev_section = None
        for row_data in wide_table:
            section = row_data["报表类型"]
            if section != prev_section:
                # 跨分类时空一行（首个分类前不空）
                if prev_section is not None:
                    row += 1
                # 写分类标题
                cell = ws_wide.cell(row=row, column=1, value=section)
                cell.font = cat_font
                cell.fill = cat_fill
                ws_wide.merge_cells(
                    start_row=row, start_column=1, end_row=row,
                    end_column=len(years) + 2)
                row += 1
                prev_section = section
            # 写数据
            ws_wide.cell(row=row, column=1, value=row_data["报表类型"])
            ws_wide.cell(row=row, column=2, value=row_data["会计科目"])
            for col_idx, year in enumerate(years, 3):
                v = row_data.get(_yl(year))
                cell = ws_wide.cell(row=row, column=col_idx)
                if v is None:
                    cell.value = "-"
                else:
                    cell.value = v
                    cell.number_format = "#,##0.00"
            row += 1
        ws_wide.column_dimensions["A"].width = 14
        ws_wide.column_dimensions["B"].width = 32
        for col_idx in range(3, len(years) + 3):
            ws_wide.column_dimensions[get_column_letter(col_idx)].width = 18

    # ===== 科目表（长格式）=====
    if long_table:
        ws_long = wb.create_sheet("科目表(长)")
        headers = ["年份", "报表类型", "会计科目", "类别", "数值"]
        for col, h in enumerate(headers, 1):
            cell = ws_long.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for i, row_data in enumerate(long_table, 2):
            # 年份可能是 tuple，统一转为字符串
            year_val = row_data["年份"]
            if isinstance(year_val, tuple):
                year_val = f"{year_val[0]}-{str(year_val[1]).zfill(2)}"
            ws_long.cell(row=i, column=1, value=year_val)
            ws_long.cell(row=i, column=2, value=row_data["报表类型"])
            ws_long.cell(row=i, column=3, value=row_data["会计科目"])
            ws_long.cell(row=i, column=4, value=row_data["类别"])
            cell = ws_long.cell(row=i, column=5, value=row_data["数值"])
            cell.number_format = "#,##0.00"
        ws_long.column_dimensions["A"].width = 12
        ws_long.column_dimensions["B"].width = 14
        ws_long.column_dimensions["C"].width = 32
        ws_long.column_dimensions["D"].width = 10
        ws_long.column_dimensions["E"].width = 20

    # ===== 期末较期初变化 Sheet =====
    file_info = {}
    for y in years:
        yd = data_by_year.get(y, {})
        if isinstance(yd, dict) and '__files__' in yd:
            file_info[y] = yd['__files__']
    _write_change_sheet(wb, data_by_year, years, file_info,
                          header_font, header_fill, cat_font, cat_fill)

    wb.save(output_path)


# ============================================================
# CLI 入口
# ============================================================

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("用法: python analyzer.py <财报文件1> [财报文件2] [财报文件3] ...")
        print("     支持传入多年财报文件，自动按文件名中的年份归类")
        sys.exit(1)

    file_list = sys.argv[1:]
    print(f"📂 待分析文件: {len(file_list)} 个")
    for f in file_list:
        year = detect_year_from_filename(f)
        print(f"  • {Path(f).name}  →  年份: {year or '未识别'}")

    data_by_year, indicators_by_year, years, trends, errors = analyze_files(file_list)
    print(f"\n✅ 成功分析 {len(years)} 年: {years}")

    if errors:
        print(f"\n⚠️ {len(errors)} 个错误:")
        for fp, err in errors:
            print(f"  • {Path(fp).name}: {err[:100]}")

    print(f"\n📈 各年指标数:")
    for y in years:
        print(f"  {y}: {len(indicators_by_year.get(y, {}))} 项")

    print(f"\n📊 趋势分析:")
    for k, v in list(trends.items())[:5]:
        print(f"  {k}: CAGR={v.get('cagr')}, YoY={[y['yoy'] for y in v.get('yoy', [])]}")

    # 生成报告
    if years:
        output = f"财务分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        generate_report(data_by_year, indicators_by_year, years, trends, output,
                        "测试公司", errors)
        print(f"\n💾 报告已生成: {output}")
