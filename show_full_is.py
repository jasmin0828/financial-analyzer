"""看利润表完整内容"""
from openpyxl import load_workbook

f = '/Users/jasmin0828/Downloads/蓉惠芯创基础资料/9、2024-2026年3月财务报表/2025年12月利润表.xlsx'
wb = load_workbook(f, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
print(f"━━━ 利润表完整内容 ({ws.max_row} 行 x {ws.max_column} 列) ━━━")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    cells = [str(c)[:25] if c is not None else '' for c in row]
    print(f"行 {i:2d}: {' | '.join(cells)}")
